import sys
import os
from PySide6.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QDialog, QLineEdit, QFormLayout, QDateEdit, QTextEdit, QTableWidget, QTableWidgetItem, QLabel, QDialogButtonBox, QComboBox, QMessageBox, QFileDialog
from PySide6.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import matplotlib.pyplot as plt
import pandas as pd
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter



# Inicializa Firebase Admin SDK
cred = credentials.Certificate("C:/Users/cop406l/Desktop/terminado/consultorio-ce70b-firebase-adminsdk-pz6po-b0e72c0be9.json")
firebase_admin.initialize_app(cred, {'databaseURL': "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-lss3s%40consultorior-37acd.iam.gserviceaccount.com"})
db = firestore.client()

class VentanaSesiones(QMainWindow):
    def __init__(self):
        super(VentanaSesiones, self).__init__()
        self.setWindowTitle("Sesiones")

        self.button_iniciar = QPushButton("INICIAR")
        self.button_iniciar.clicked.connect(self.abrir_ventana_opciones)

        self.button_exportar_excel = QPushButton("Exportar a Excel")
        self.button_exportar_excel.clicked.connect(self.exportar_a_excel)

        self.rut_busqueda = QLineEdit()
        self.rut_busqueda.setPlaceholderText("Ingrese RUT para buscar")
        self.rut_busqueda.returnPressed.connect(self.buscar_datos)

        self.figure1 = plt.Figure()
        self.canvas1 = FigureCanvas(self.figure1)
        self.ax1 = self.figure1.add_subplot(111)

        self.figure2 = plt.Figure()
        self.canvas2 = FigureCanvas(self.figure2)
        self.ax2 = self.figure2.add_subplot(111)

        self.button_abrir_pacientes = QPushButton("Abrir Pacientes")
        self.button_abrir_pacientes.clicked.connect(self.abrir_ventana_pacientes)

        layout = QVBoxLayout()
        layout.addWidget(self.button_iniciar)
        layout.addWidget(self.rut_busqueda)
        layout.addWidget(self.canvas1)
        layout.addWidget(self.canvas2)
        layout.addWidget(self.button_exportar_excel)  # Agregar botón de exportar a Excel
        layout.addWidget(self.button_abrir_pacientes)  # Agregar botón para abrir ventana de pacientes

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # Configurar tamaño de ventana
        self.resize(800, 600)

        # Establecer estilos
        self.setStyleSheet("""
            QPushButton {
                background-color: blue;
                border-radius: 10px;
                color: white;
            }
        """)

        # Mostrar la ventana de sesiones al iniciar
        self.buscar_datos()

    def abrir_ventana_opciones(self):
        self.ventana_opciones = VentanaOpciones()
        self.ventana_opciones.show()
        
    def abrir_ventana_pacientes(self):
        self.ventana_pacientes = VentanaPacientes()
        self.ventana_pacientes.setStyleSheet(self.styleSheet())  # Aplicar el mismo estilo de la ventana de sesiones a la de pacientes
        self.ventana_pacientes.show()

    def buscar_datos(self):
        rut = self.rut_busqueda.text()
        docs = db.collection('sesiones').stream()
        datos = [doc.to_dict() for doc in docs if doc.to_dict().get('rut') == rut]
        self.visualizar_grafico_dolor(datos)
        self.visualizar_grafico_fuerza(datos)

    def visualizar_grafico_dolor(self, datos):
        self.ax1.clear()  # Limpiar el gráfico antes de actualizar

        sesiones = [int(dato.get('sesion', 1)) for dato in datos]
        dolor_reposo = [int(dato.get('dolor_reposo', 0)) for dato in datos]
        dolor_movimiento = [int(dato.get('dolor_movimiento', 0)) for dato in datos]

        if not sesiones:  # Verificar si la lista de sesiones está vacía
            # Mostrar un mensaje de advertencia al usuario
           
            return

        # Graficar barras de dolor en reposo y en movimiento
        self.ax1.bar(sesiones, dolor_reposo, width=0.4, align='center', label='Dolor en Reposo')
        self.ax1.bar([x + 0.4 for x in sesiones], dolor_movimiento, width=0.4, align='center', label='Dolor en Movimiento')

        # Configuraciones del gráfico
        self.ax1.set_xlabel('Número de Sesión')
        self.ax1.set_ylabel('Dolor')
        self.ax1.set_title('Dolor en Reposo y en Movimiento por Sesión')
        self.ax1.set_xticks(sesiones)
        self.ax1.legend()

        # Actualizar el lienzo del gráfico
        self.canvas1.draw()

    def visualizar_grafico_fuerza(self, datos):
        self.ax2.clear()  # Limpiar el gráfico antes de actualizar

        sesiones = [int(dato.get('sesion', 1)) for dato in datos]
        fuerza_muscular = [int(dato.get('fuerza_muscular', 0)) for dato in datos]

        if not sesiones:  # Verificar si la lista de sesiones está vacía
            # Mostrar un mensaje de advertencia al usuario
            QMessageBox.warning(None, "Visualizar Gráfico", "No hay datos disponibles para graficar.")
            return

        # Graficar barras de fuerza muscular
        self.ax2.bar(sesiones, fuerza_muscular, width=0.4, align='center', label='Fuerza Muscular')

        # Configuraciones del gráfico
        self.ax2.set_xlabel('Número de Sesión')
        self.ax2.set_ylabel('Fuerza Muscular')
        self.ax2.set_title('Fuerza Muscular por Sesión')
        self.ax2.set_xticks(sesiones)
        self.ax2.legend()

        # Actualizar el lienzo del gráfico
        self.canvas2.draw()

    # Dentro del método exportar_a_excel de la clase VentanaSesiones

    def exportar_a_excel(self):
        rut = self.rut_busqueda.text()
        docs = db.collection('sesiones').stream()
        datos_sesiones = [doc.to_dict() for doc in docs if doc.to_dict().get('rut') == rut]

        if not datos_sesiones:
            QMessageBox.warning(self, "Exportar a Excel", "No hay datos disponibles para exportar.")
            return

        # Ordenar los datos por el número de sesión
        datos_sesiones_ordenados = sorted(datos_sesiones, key=lambda x: int(x.get('sesion', 0)))

        # Crear un nuevo libro de Excel y seleccionar la hoja activa
        wb = Workbook()
        ws = wb.active

        # Definir el alto de las filas en 60 píxeles
        for i in range(1, len(datos_sesiones_ordenados) + 2):
            ws.row_dimensions[i].height = 60 / 1.3  # Convertir píxeles a unidades de alto de fila en Excel

        # Ajustar el ancho de las columnas
        column_widths = [130, 220, 138, 138, 138]  # Anchos de las columnas en pixeles
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width / 7.5  # Convertir pixeles a unidades de ancho de columna en Excel

        # Agregar texto "RESPONSABLE" en la celda B1
        ws['B1'] = "RESPONSABLE"

        # Agregar encabezados
        encabezados = ["Sesión", "RUT", "Evolución", "Fisioterapia", "Kinesiterapia"]
        ws.append(encabezados)

        # Agregar datos de sesiones
        for dato in datos_sesiones_ordenados:
            sesion = dato.get("sesion", "")
            rut = dato.get("rut", "")
            evolucion = dato.get("evolucion", "")
            fisioterapia = dato.get("fisioterapia", "")
            kinesiterapia = dato.get("kinesiterapia", "")
            ws.append([sesion, rut, evolucion, fisioterapia, kinesiterapia])

        # Ajustar la altura de las filas y el ancho de las columnas
        for row in ws.iter_rows(min_row=2, max_row=len(datos_sesiones_ordenados) + 2, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')  # Ajustar el texto para que se ajuste dentro de las celdas
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))  # Agregar bordes internos y contorno

        # Ajustar márgenes de la hoja
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # Guardar el archivo Excel
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getSaveFileName(self, "Guardar como", "", "Excel Files (*.xlsx)", options=options)
        if file_name:
            wb.save(file_name)
            QMessageBox.information(self, "Exportar a Excel", f"Los datos se han exportado a '{file_name}' correctamente.")
class VentanaPacientes(QMainWindow):
    def __init__(self):
        super(VentanaPacientes, self).__init__()
        self.setWindowTitle("Pacientes")

        self.button_nuevo_paciente = QPushButton("Nuevo Paciente")
        self.button_nuevo_paciente.clicked.connect(self.abrir_formulario_paciente)

        self.table_pacientes = QTableWidget()
        self.table_pacientes.setColumnCount(13)
        self.table_pacientes.setHorizontalHeaderLabels(["Nombre", "Edad", "Teléfono", "RUT", "Fecha", "Médico Tratante", "Cirugía", "Diagnóstico", "Previsión", "Dirección", "Correo Electrónico", "Motivo de Consulta", "Sesiones"])
        self.actualizar_tabla_pacientes()

        self.button_abrir_sesiones = QPushButton("Abrir Sesiones")
        self.button_abrir_sesiones.clicked.connect(self.abrir_ventana_sesiones)

        self.rut_busqueda = QLineEdit()
        self.rut_busqueda.setPlaceholderText("Buscar por RUT")
        self.rut_busqueda.returnPressed.connect(self.buscar_paciente)

        layout = QVBoxLayout()
        layout.addWidget(self.button_nuevo_paciente)
        layout.addWidget(self.rut_busqueda)
        layout.addWidget(self.table_pacientes)
        layout.addWidget(self.button_abrir_sesiones)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        self.resize(800, 600)

        self.setStyleSheet("""
            QPushButton {
                background-color: blue;
                border-radius: 10px;
                color: white;
            }
        """)

    def abrir_ventana_sesiones(self):
        self.ventana_sesiones = VentanaSesiones()
        self.ventana_sesiones.show()


    def abrir_formulario_paciente(self):
        self.formulario_paciente = FormularioPaciente()
        self.formulario_paciente.show()


    def actualizar_tabla_pacientes(self):
        pacientes = self.recuperar_pacientes()
        self.table_pacientes.setRowCount(len(pacientes))
        for i, paciente in enumerate(pacientes):
            self.table_pacientes.setItem(i, 0, QTableWidgetItem(paciente.get("nombre", "")))
            self.table_pacientes.setItem(i, 1, QTableWidgetItem(paciente.get("edad", "")))
            self.table_pacientes.setItem(i, 2, QTableWidgetItem(paciente.get("telefono", "")))
            self.table_pacientes.setItem(i, 3, QTableWidgetItem(paciente.get("rut", "")))
            self.table_pacientes.setItem(i, 4, QTableWidgetItem(paciente.get("fecha", "")))
            self.table_pacientes.setItem(i, 5, QTableWidgetItem(paciente.get("medicoTratante", "")))
            self.table_pacientes.setItem(i, 6, QTableWidgetItem(paciente.get("cirugia", "")))
            self.table_pacientes.setItem(i, 7, QTableWidgetItem(paciente.get("diagnostico", "")))
            self.table_pacientes.setItem(i, 8, QTableWidgetItem(paciente.get("prevision", "")))
            self.table_pacientes.setItem(i, 9, QTableWidgetItem(paciente.get("direccion", "")))
            self.table_pacientes.setItem(i, 10, QTableWidgetItem(paciente.get("correo", "")))
            item = self.table_pacientes.item(i, 11)
            if item is not None:
                item.setText(paciente.get("motivo_consulta", ""))

            button_nueva_sesion = QPushButton("Nueva Sesión")
            button_nueva_sesion.clicked.connect(lambda _, rut=paciente.get("rut", ""): self.nueva_sesion(rut))  # Conectar el clic del botón a una función
            self.table_pacientes.setCellWidget(i, 12, button_nueva_sesion)  # Establecer el botón en la celda correspondiente

    def nueva_sesion(self, rut):
        # Esta función maneja lo que sucede cuando se hace clic en el botón "Nueva Sesión"
        self.button_nueva_sesion = QPushButton("Nueva Sesión")
        self.button_nueva_sesion.clicked.connect(self.abrir_formulario_sesion)
        # Puedes abrir un formulario para ingresar los detalles de la nueva sesión o realizar cualquier otra acción deseada
        pass  # Aquí puedes implementar la lógica para abrir el formulario de nueva sesión o realizar otras acciones
                
    def recuperar_pacientes(self):
        pacientes = []
        docs = db.collection("pacientes").stream()
        for doc in docs:
            paciente = doc.to_dict()
            # Asegurarse de que la clave 'sesiones' exista en el diccionario del paciente
            paciente.setdefault('sesiones', [])
            pacientes.append(paciente)
        return pacientes

    def buscar_paciente(self):
        rut = self.rut_busqueda.text()
        rows = self.table_pacientes.rowCount()
        for i in range(rows):
            item = self.table_pacientes.item(i, 3)
            if item and item.text() == rut:
                if i != 0:
                    self.mover_fila_al_principio(i)
                return
        QMessageBox.warning(self, "Buscar Paciente", "No se encontró ningún paciente con el RUT proporcionado.")

    def mover_fila_al_principio(self, row):
        column_count = self.table_pacientes.columnCount()
        new_row = []
        for col in range(column_count):
            item = self.table_pacientes.takeItem(row, col)
            if item:
                new_row.append(item)
        self.table_pacientes.insertRow(0)
        for col, item in enumerate(new_row):
            self.table_pacientes.setItem(0, col, item)
        self.table_pacientes.removeRow(row)

class VentanaOpciones(QDialog):
    def __init__(self):
        super(VentanaOpciones, self).__init__()
        self.setWindowTitle("Opciones")

        self.button_nueva_sesion = QPushButton("Nueva Sesión")
        self.button_nueva_sesion.clicked.connect(self.abrir_formulario_sesion)

        self.button_nuevo_cliente = QPushButton("Nuevo Cliente")
        self.button_nuevo_cliente.clicked.connect(self.abrir_formulario_paciente)

        layout = QVBoxLayout()
        layout.addWidget(self.button_nueva_sesion)
        layout.addWidget(self.button_nuevo_cliente)

        container = QWidget()
        container.setLayout(layout)
        self.setLayout(layout)

    def abrir_formulario_sesion(self):
        self.formulario_sesion = FormularioSesion()
        self.formulario_sesion.show()

    def abrir_formulario_paciente(self):
        self.formulario_paciente = FormularioPaciente()
        self.formulario_paciente.show()

class FormularioPaciente(QDialog):
    def __init__(self):
        super(FormularioPaciente, self).__init__()
        self.setWindowTitle("Nuevo Paciente")

        self.nombre = QLineEdit()
        self.edad = QLineEdit()
        self.telefono = QLineEdit()
        self.rut = QLineEdit()
        self.fechaNacimiento = QDateEdit()
        self.medicoTratante = QLineEdit()
        self.cirugia = QLineEdit()
        self.prevision = QLineEdit()
        self.direccion = QLineEdit()
        self.correo = QLineEdit()
        self.motivo_consulta = QTextEdit()

        layout = QFormLayout(self)
        layout.addRow(QLabel("Formulario de Paciente"))
        layout.addRow("Nombre", self.nombre)
        layout.addRow("Edad", self.edad)
        layout.addRow("Teléfono", self.telefono)
        layout.addRow("RUT", self.rut)
        layout.addRow("Fecha de Nacimiento", self.fechaNacimiento)
        layout.addRow("Médico Tratante", self.medicoTratante)
        layout.addRow("Cirugía", self.cirugia)
        layout.addRow("Previsión", self.prevision)
        layout.addRow("Dirección", self.direccion)
        layout.addRow("Correo Electrónico", self.correo)
        layout.addRow("motivo de consulta", self.motivo_consulta)

        buttonBox = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttonBox.accepted.connect(self.guardar_datos)
        buttonBox.rejected.connect(self.reject)
        layout.addRow(buttonBox)

    def guardar_datos(self):
        data = {
            "nombre": self.nombre.text(),
            "edad": self.edad.text(),
            "telefono": self.telefono.text(),
            "rut": self.rut.text(),
            "fecha": self.fechaNacimiento.date().toString(Qt.ISODate),
            "medicoTratante": self.medicoTratante.text(),
            "cirugia": self.cirugia.text(),
            "prevision": self.prevision.text(),
            "direccion": self.direccion.text(),
            "correo": self.correo.text(),
            "motivo de consulta": self.motivo_consulta.toPlainText()
        }

        try:
            db.collection("pacientes").add(data)
            QMessageBox.information(self, "Guardar Datos", "Los datos del paciente se han guardado correctamente.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudieron guardar los datos del paciente: {str(e)}")

class FormularioSesion(QDialog):
    def __init__(self):
        super(FormularioSesion, self).__init__()
        self.setWindowTitle("Nueva Sesión")

       
        self.rut = QLineEdit()
        self.Ncesion = QLineEdit()
        self.fecha = QDateEdit()
        self.medico = QLineEdit()
        self.observaciones = QTextEdit()
        self.evolucion = QTextEdit()
        self.examen_fisico = QComboBox()
        self.examen_fisico.addItems(["Normal", "Anormal"])
        self.evaluacion_muscular = QComboBox()
        self.evaluacion_muscular.addItems(["Normal", "Anormal"])
        self.medicion_articular = QLineEdit()
        self.fuerza_muscular = QComboBox()
        self.fuerza_muscular.addItems(["0", "1", "2", "3", "4", "5"])
        self.dolor_reposo = QComboBox()
        self.dolor_reposo.addItems(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"])
        self.dolor_movimiento = QComboBox()
        self.dolor_movimiento.addItems(["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"])
        self.actitud_funcional = QComboBox()
        self.actitud_funcional.addItems(["Normal", "Anormal"])

        layout = QFormLayout(self)
        layout.addRow(QLabel("Formulario de Sesión"))
        layout.addRow("RUT", self.rut)
        layout.addRow("N° de Sesión", self.Ncesion)
        layout.addRow("Fecha", self.fecha)
        layout.addRow("Médico", self.medico)
        layout.addRow("Observaciones", self.observaciones)
        layout.addRow("Evolución", self.evolucion)
        layout.addRow("Examen Físico", self.examen_fisico)
        layout.addRow("Evaluación Muscular", self.evaluacion_muscular)
        layout.addRow("Medición Articular", self.medicion_articular)
        layout.addRow("Fuerza Muscular", self.fuerza_muscular)
        layout.addRow("Dolor en Reposo", self.dolor_reposo)
        layout.addRow("Dolor en Movimiento", self.dolor_movimiento)
        layout.addRow("Actitud Funcional", self.actitud_funcional)

        buttonBox = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttonBox.accepted.connect(self.guardar_datos)
        buttonBox.rejected.connect(self.reject)
        layout.addRow(buttonBox)

    def guardar_datos(self):
        data = {
            "rut": self.rut.text(),
            "sesion": self.Ncesion.text(),
            "fecha": self.fecha.date().toString(Qt.ISODate),
            "medico": self.medico.text(),
            "observaciones": self.observaciones.toPlainText(),
            "evolucion": self.evolucion.toPlainText(),
            "examen_fisico": self.examen_fisico.currentText(),
            "evaluacion_muscular": self.evaluacion_muscular.currentText(),
            "medicion_articular": self.medicion_articular.text(),
            "fuerza_muscular": self.fuerza_muscular.currentText(),
            "dolor_reposo": self.dolor_reposo.currentText(),
            "dolor_movimiento": self.dolor_movimiento.currentText(),
            "actitud_funcional": self.actitud_funcional.currentText()
        }

        try:
            db.collection("sesiones").add(data)
            QMessageBox.information(self, "Guardar Datos", "Los datos de la sesión se han guardado correctamente.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudieron guardar los datos de la sesión: {str(e)}")

if __name__ == "__main__":
    app = QApplication([])
    window = VentanaSesiones()
    window.show()
    sys.exit(app.exec_())
